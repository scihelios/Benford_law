from mpl_toolkits import mplot3d
import random
import math as m
import openpyxl
import itertools 
import time

import numpy as np
import matplotlib.pyplot as plt
import tkinter as tk

from tkinter import *
from tkinter import ttk
from tkinter import scrolledtext
from tkinter import Menu
from tkinter.ttk import Progressbar
from tkinter.ttk import *

from fpdf import FPDF
from matplotlib.backends.backend_pdf import PdfPages



#Je m'excuse en avance pour toutes les fautes d'orthographes dans les commentaires.



#creer une vecteur de R^9 qui présente le graphe de la distrubition théorique
#@jit
def Benford(N):
	return np.array([N*(m.log10(1+1/i)) for i in range(1,10)])

#une fonction qui permet de calculer la distance entre deux graphe
#@jit 
def distance(f,g):
	#sommer la distance pour chaque occurence de chiffre
	g1=g+np.ones(9)
	f1=f+np.ones(9)
	return max(np.linalg.norm((f1-g1)/g1,1),np.linalg.norm((g1-f1)/f1,1))

#une fonction qui produit une série de variables aléaatoires indpendantes et invariantes qui suivent loi de benford 
def valmaker(N,E):
	#créer l'ensemble des résultat d'un essai
	return np.random.multinomial(N, [m.log10(1+(1/i)) for i in range(1,10)], size=E)


#fonction pour rendre la fenetre en mode plein écran
class FullScreenApp(object):
	def __init__(self, master, **kwargs):
		self.master=master
		pad=3
		self._geom='200x200+0+0'
		master.geometry("{0}x{1}+0+0".format(
			master.winfo_screenwidth()-pad, master.winfo_screenheight()-pad))
		master.bind('<Escape>',self.toggle_geom)
	def toggle_geom(self,event):
		geom=self.master.winfo_geometry()
		print(geom,self._geom)
		self.master.geometry(self._geom)
		self._geom=geom


#prend un fichier CSV et enregistre dans une liste les occurance de chaque chiffre et dans une autre liste les nombre eux meme 
def scrapcvs(fichier):
	#initialisation
	tabnum=[0,0,0,0,0,0,0,0,0]
	nombre=[]
	
	#recuperer les occurance et chiffre à l'aide de compteur (faut lire le code et essayer de comprendre)
	with open(fichier,"r") as file:
		for ligne in file:
			if ligne[0]=="":
				return tabnum
			inomb=''
			k=0
			j=0
			for c in ligne:
				if c in '0123456789':
					if k==0 or j==1:
						if c=='0':
							j=1
						else:
							tabnum[int(c)-1] += 1
							j=0
							k=1
					inomb += c	
				else:
					if c=='.':
						inomb += c
						k=2
					else:
						if k==1:
							nombre += [m.trunc(float(inomb))+1]
							k=0
							j=0
							inomb=''
	
	return [tabnum,nombre]


#la fonction de bouton tester ; permet de prévisualiser le graphe
def test():
	#recuperer les données dèja saisies
	a=texta.get()
	b=textb.get()
	c=textc.get()
	d=textd.get()
	e=texte.get()

	#créer l'objet fig où les figures sont enregistré avec ax[0] et ax[1] les sous-graphs
	fig=plt.figure( figsize=(12, 9))
	ax = fig.subplots(2, sharex=True, sharey=True)
	
	#prendre les données deja traitéés
	data=scrapcvs(d)		
	x1=data[0]
	x2=data[1]
	y=[1,2,3,4,5,6,7,8,9]
	
	#mettre l'histogramme théorique et emperique cote à cote
	plt.xticks(y,['1','2','3','4','5','6','7','8','9'])
	ax[0].bar([i-0.15 for i in range(1,10)],[len(x2)*(m.log10(1+1/i)) for i in range(1,10)],width=0.3)
	ax[0].bar([i+0.15for i in range(1,10)],x1,width=0.3)
	      
	      
	#arranger les nombres de façon croissante
	xsort= sorted(x2)
	cont=0
	y2=[-1]

	#determiner le nombre de point de données qui se trouve dans chaque ordre de grandeur 
	ordgrand=len(str(xsort[0]))
	ordre=[ordgrand]
	for i in xsort:
		ordgrand1=len(str(i))
		y2[cont] += 1
		if ordgrand1!= ordgrand:
			cont += 1
			y2+= [0]
			ordgrand=ordgrand1
			ordre += [ordgrand1]

	#tracer le graph des ordres de grandeurs
	plt.xticks([i for i in ordre],['10^'+str(i) for i in ordre])		
	ax[1].bar([i for i in ordre],y2,width=0.4)
	
	plt.show()		

#créer les pdfs qui contiennent le rapport ecrit et la figure de comparaison
def imprimer():
	global window
	global texta
	global textb
	global textc
	global textd
	global texte
	
	#fonctionalité désactivée lors du testing
	'''bar = Progressbar(window, mode='determinate', length=500, style='black.Horizontal.TProgressbar')
	bar['value'] = 1
	bar.place(relx=0.5, rely=0.9, anchor=CENTER)
	window.update_idletasks()
	time.sleep(0.5)
	window.update_idletasks()
	for j in range(52):
		window.update_idletasks()
		bar['value'] = j*2
		time.sleep(random.random()/50)
	bar.destroy()'''


	#recuperer les données déja saisies
	a=texta.get()
	b=textb.get()
	c=textc.get()
	d=textd.get()
	e=texte.get()

	#créer l'objet fig où les figures sont enregistré avec ax[0] et ax[1] les sous-graphs
	fig=plt.figure( figsize=(12, 9))
	ax = fig.subplots(3, sharex=False, sharey=False)
	
	#prendre les donnée deja traité
	data=scrapcvs(d)		
	x1=data[0]
	x2=data[1]
	y=[1,2,3,4,5,6,7,8,9]
	
	#mettre l'histogramme théorique et emperique cote à cote
	plt.sca(ax[0])
	plt.xticks(y,['1','2','3','4','5','6','7','8','9'])
	ax[0].bar([i-0.15 for i in range(1,10)],[len(x2)*(m.log10(1+1/i)) for i in range(1,10)],width=0.3)
	ax[0].bar([i+0.15for i in range(1,10)],x1,width=0.3)
	
	#arranger les nombres de façon croissante
	xsort= sorted(x2)
	cont=0
	y2=[-1]

	#determiner le nombre de point de données qui se trouve dans chaque ordre de grandeur 
	ordgrand=len(str(xsort[0]))
	ordre=[ordgrand]
	for i in xsort:
		ordgrand1=len(str(i))
		y2[cont] += 1
		if ordgrand1!= ordgrand:
			cont += 1
			y2+= [0]
			ordgrand=ordgrand1
			ordre += [ordgrand1]

	#tracer le graph des ordres de grandeurs
	plt.sca(ax[1])
	plt.xticks([i for i in ordre],['10^'+str(i) for i in ordre])	
	ax[1].bar([i for i in ordre],y2,width=0.4)

	N=len(x2)
	Dmax=m.trunc(distance(x1,Benford(len(x2)))+1)
	exprm=valmaker(N,10000)
	gth=Benford(N)
	X=[Dmax*(i/200) for i in range(200) ]
	Y=[0 for i in range(200)]
	for i in exprm:
		dg=distance(i,gth)
		k=0
		while (k<200) and (dg>X[k]):
			Y[k] += (1/10000)
			k+=1
	ax[2].plot(X,Y)
	disr=distance(x1,Benford(len(x2)))
	gr=[[disr,disr],[0,1]]
	ax[2].plot(gr[0],gr[1])
	plt.show()

	#exporter la figure vers un pdf
	fig.savefig('figure de '+e+'.pdf')

	#creer un nouveau pdf
	mpdf=FPDF('P', 'mm', 'Letter')
	mpdf.add_page()
	mpdf.set_font('helvetica', '',13)
	
	#definir la couleur des cells si il y a remplissage
	

	#mettre dans le pdf les infos déja saisie
	mpdf.cell(200, 8,e, align='C',border=0 ,ln=1)
	mpdf.cell(100, 10,'', ln=1)
	mpdf.cell(180, 8,'Date : '+a, border=1 , ln=1)
	mpdf.cell(100, 4,'',  ln=1)
	mpdf.cell(180, 8,'Nom de fichier : '+b, border=1, ln=1)
	mpdf.cell(100, 4,'',  ln=1)
	mpdf.cell(180, 8,'Réference de test : '+c,border=1,ln=1)
	mpdf.cell(100, 4,'',  ln=1)
	mpdf.multi_cell(180, 8,'URL du fichier CSV : '+d,border=1, ln=1)
	mpdf.cell(100, 8,'',  ln=1)
	mpdf.multi_cell(200, 8, 'infos de bases pour ce test (pour comprendre les résultats dans leur contexte) ;  ',border=0, ln=1)
	mpdf.cell(100, 4,'',  ln=1)
	mpdf.multi_cell(150, 8, "       1- Cardinal de l'échantillion : |D| = "+str(len(x2)),border=0, ln=1)
	mpdf.multi_cell(150, 8, "       2- pour un tel cadinal une distance maximal atteinte est de : "+str(m.trunc(X[sum([1 for i in range(199) if Y[i]>0])]*1000)/1000),border=0, ln=1)
	mpdf.multi_cell(150, 8, "       3- la distance du graphe des données fournis est = "+str(m.trunc(disr*1000)/1000),border=0, ln=1)
	mpdf.cell(100, 12,'',  ln=1)
	#verification si on est bien dans la marge acceptable ou pas grace au test statistique
	avis = X[sum([1 for i in range(200) if Y[i]>0.005])] < disr
	if avis:
		avis='inacceptable'
	else:
		avis='acceptable'

	mpdf.multi_cell(150, 8, 'Après analyse sattistique des données fournis on conclut que ;  ',border=0, ln=1)
	mpdf.multi_cell(150, 8, "         ce résultat est "+avis+'.',border=0, ln=1)
	if  avis=='inacceptable':
		mpdf.cell(100, 4,'',  ln=1)
		mpdf.multi_cell(180, 8, "         les données ne suivent pas la loi de Benford avec une certitude de 99.5%.",border=0, ln=1)
		mpdf.cell(100, 12,'',  ln=1)
		mpdf.set_text_color(205,0,0)
		mpdf.multi_cell(180, 8,"Meme si ce test confirme vos souspicion de fraude, il ne doit pas etre le seul outil à utiliser pour accuser les parties concernées.",border=1,ln=1)
		mpdf.set_text_color(0,0,0)
	if  avis=='acceptable':
		mpdf.cell(100, 4,'',  ln=1)
		mpdf.multi_cell(180, 8, "         les données suivent la loi de Benford dans la marge de confiance prédeterminé.",border=0, ln=1)
	mpdf.cell(100, 20,'',  ln=1)
	mpdf.multi_cell(200, 8, 'figure de '+e+'.pdf', align='C',border=0, ln=1)
	mpdf.cell(100, 20,'',  ln=1)
	mpdf.cell(200, 8,'-1-', align='C', ln=1)

	#exporter le pdf
	mpdf.output(e+'.pdf')

#réinitialiser les zones de text	
def reset1():
	texta.delete(0, tk.END)
	textb.delete(0, tk.END)
	textc.delete(0, tk.END)
	textd.delete(0, tk.END)
	texte.delete(0, tk.END)
	
#cette fonction permet de creer une fenetre aprés une verification réussite
def clicked1():
	global window
	global texta
	global textb
	global textc
	global textd
	global texte	

	#fonctionalité désactivée lors du testing

	'''bar = Progressbar(window, mode='determinate', length=500, style='black.Horizontal.TProgressbar')
	bar['value'] = 1
	bar.place(relx=0.5, rely=0.9, anchor=CENTER)
	window.update_idletasks()
	time.sleep(0.5)
	window.update_idletasks()
	for j in range(52):
		window.update_idletasks()
		bar['value'] = j*2
		time.sleep(random.random()/50)
	bar.destroy()'''


	#creer une nouvelle fenetre
	window.destroy()
	window = Tk()
	window.title("nouvelle composition")
	window.geometry('600x600')
	window.option_add("*tearOff", False)
	
	FullScreenApp(window)

	#créer les zone de text
	texta = Entry(window,width=20,font=("Arial Bold", 15))
	textb = Entry(window,width=20,font=("Arial Bold", 15))
	textc = Entry(window,width=20,font=("Arial Bold", 15))
	textd = Entry(window,width=30,font=("Arial Bold", 15))
	texte = Entry(window,width=20,font=("Arial Bold", 15))

	lbld = Label(window, text="-----------------Nouveau test-------------", font=("Arial Bold", 30))
	lbld.place(relx=0.5, rely=0.1, anchor=CENTER)

	lbln = Label(window, text="Titre  :", font=("Arial Bold", 15))
	lbln.place(relx=0.2, rely=0.3, anchor=W)
	texte.place(relx=0.7, rely=0.3, anchor=W)

	lblp = Label(window, text="Date : ",font=("Arial Bold", 15))
	lblp.place(relx=0.2, rely=0.4, anchor=W)
	texta.place(relx=0.7, rely=0.4, anchor=W)
	
	lbla = Label(window, text="Reference du test : ",font=("Arial Bold", 15))
	lbla.place(relx=0.2, rely=0.5, anchor=W)
	textb.place(relx=0.7, rely=0.5, anchor=W)
		
	lblb = Label(window, text="Nom du fichier : ",font=("Arial Bold", 15))
	lblb.place(relx=0.2, rely=0.6, anchor=W)
	textc.place(relx=0.7, rely=0.6, anchor=W)
	
	lblc = Label(window, text="Adresse du fichier : ",font=("Arial Bold", 15))
	lblc.place(relx=0.2, rely=0.7, anchor=W)
	textd.place(relx=0.7, rely=0.7, anchor=W)
	
	#creer les boutons
	btncomm1 = Button(window, text="TESTER", command=test)
	btncomm1.place(relx=0.8, rely=0.875, anchor=CENTER)

	btncomms = Button(window, text="IMPRIMER", command=imprimer)
	btncomms.place(relx=0.5, rely=0.875, anchor=CENTER)
	
	btncomm2 = Button(window, text="ANNULER", command=reset1)
	btncomm2.place(relx=0.2, rely=0.875, anchor=CENTER)


#fonction pour verifier si l'utilisateur est authorisé à l'utilisation du programme (désactivé lors du testing)
def clickverif():
	global window

	#prendre les infos de authentification
	ide=txta.get()
	mp=txtb.get()
	
	#verification des données saisie (désactivé lors du testing d'où le 1==1)
	if 1==1:
		verify.destroy()
		window = Tk()

		window.title("test")

		window.geometry('600x600')
		window.option_add("*tearOff", False)
		
		
		lbl1 = Label(window, text="veuillez suivre les instructions correctement :" )
		lbl1.place(relx=0.5, rely=0.3, anchor=CENTER)
		lbl1 = Label(window, text="1- mettre le l'adresse du fichier correctement en utilisant un double backslash (\\\\)." )
		lbl1.place(relx=0.5, rely=0.35, anchor=CENTER)
		lbl1 = Label(window, text="exemple : C:\\\\Users\\\\ahmed\\\\gdptest.csv ." )
		lbl1.place(relx=0.5, rely=0.4, anchor=CENTER)
		lbl1 = Label(window, text="2- le fichier doit etre en format CVS et arrangée comme dans l'exemple indiqué ." )
		lbl1.place(relx=0.5, rely=0.45, anchor=CENTER)
		lbl2 = Label(window, text="cliquez sur COMMENCER pour continuer ")
		lbl2.place(relx=0.5, rely=0.5, anchor=CENTER)

		btncomm = Button(window, text="COMMENCER", command=clicked1)
		btncomm.place(relx=0.5, rely=0.6, anchor=CENTER)
		
	
	else :
		txta.delete(0, tk.END)
		txtb.delete(0, tk.END)	
		lblm = Label(verify, text="mot de passe ou identifiant incorrecte")
		lblm.grid(column=0, row=2)

#create the window and set title
verify = Tk()
verify.title("verify")

# input dimensons of the wndow
verify.geometry('350x150')

# no idea what this do ! 
verify.option_add("*tearOff", False)


#create buttons for identification
txta = Entry(verify,width=10)
txtb = Entry(verify,show='*',width=10)
txta.grid(column=1, row=0)
txtb.grid(column=1, row=1)
lbli = Label(verify, text="identifiant : " )
lbli.grid(column=0, row=0)
lblm = Label(verify, text="mot de passe :  ")
lblm.grid(column=0, row=1)
btncon = Button(verify, text="Connecter", command=clickverif)
btncon.grid(column=1, row=3)

verify.mainloop()